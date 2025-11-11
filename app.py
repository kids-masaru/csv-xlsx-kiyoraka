import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import os

# ページ設定
st.set_page_config(
    page_title="CSV to Excel Converter",
    page_icon="📄",
    layout="centered"
)

# カスタムCSS
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    .main {
        padding: 3rem 2rem;
        background-color: #ffffff;
    }
    
    /* カラーパレット: #2D3748(ダークグレー), #4A90E2(ソフトブルー), #E8EDF2(ライトグレー) */
    
    .header-container {
        text-align: center;
        margin-bottom: 3rem;
    }
    
    .app-title {
        color: #2D3748;
        font-size: 2.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        letter-spacing: -0.5px;
    }
    
    .app-subtitle {
        color: #718096;
        font-size: 1.1rem;
        font-weight: 400;
    }
    
    .section-title {
        color: #2D3748;
        font-size: 1.3rem;
        font-weight: 600;
        margin: 2rem 0 1rem 0;
    }
    
    .upload-area {
        border: 2px dashed #4A90E2;
        border-radius: 12px;
        padding: 2.5rem;
        text-align: center;
        background-color: #F7FAFC;
        margin: 1.5rem 0;
        transition: all 0.3s;
    }
    
    .upload-area:hover {
        background-color: #EDF2F7;
        border-color: #3182CE;
    }
    
    .info-box {
        background-color: #E8EDF2;
        border-left: 4px solid #4A90E2;
        border-radius: 8px;
        padding: 1rem 1.5rem;
        margin: 1rem 0;
        color: #2D3748;
    }
    
    .success-box {
        background-color: #E8F5E9;
        border-left: 4px solid #66BB6A;
        border-radius: 8px;
        padding: 1rem 1.5rem;
        margin: 1rem 0;
        color: #2D3748;
    }
    
    .stButton>button {
        width: 100%;
        background-color: #4A90E2;
        color: white;
        padding: 0.875rem 1.5rem;
        font-size: 1rem;
        font-weight: 600;
        border-radius: 8px;
        border: none;
        transition: all 0.3s;
        letter-spacing: 0.3px;
    }
    
    .stButton>button:hover {
        background-color: #3182CE;
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(74, 144, 226, 0.3);
    }
    
    .instruction-list {
        background-color: #F7FAFC;
        border-radius: 12px;
        padding: 2rem;
        margin: 2rem 0;
    }
    
    .instruction-item {
        display: flex;
        align-items: start;
        margin: 1.2rem 0;
        color: #2D3748;
    }
    
    .step-number {
        background-color: #4A90E2;
        color: white;
        width: 32px;
        height: 32px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 600;
        margin-right: 1rem;
        flex-shrink: 0;
    }
    
    .step-text {
        padding-top: 4px;
        font-size: 1rem;
        line-height: 1.6;
    }
    
    /* ファイルアップローダーのカスタマイズ */
    .stFileUploader {
        background-color: transparent;
    }
    
    [data-testid="stFileUploadDropzone"] {
        background-color: #F7FAFC;
        border: 2px dashed #4A90E2;
        border-radius: 12px;
    }
    
    /* データフレームのスタイル */
    .stDataFrame {
        border: 1px solid #E8EDF2;
        border-radius: 8px;
        overflow: hidden;
    }
    
    hr {
        border: none;
        border-top: 1px solid #E8EDF2;
        margin: 2rem 0;
    }
    
    .footer {
        text-align: center;
        color: #A0AEC0;
        font-size: 0.9rem;
        margin-top: 3rem;
    }
    </style>
""", unsafe_allow_html=True)

# ヘッダー
st.markdown("""
    <div class="header-container">
        <div class="app-title">CSV to Excel Converter</div>
        <div class="app-subtitle">CSVデータをExcelテンプレートに簡単変換</div>
    </div>
""", unsafe_allow_html=True)

# テンプレートファイルの確認
template_path = "template.xlsx"
if not os.path.exists(template_path):
    st.error("template.xlsx が見つかりません。同じディレクトリに配置してください。")
    st.stop()

# ファイルアップロード
st.markdown('<div class="section-title">CSVファイルをアップロード</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "ファイルを選択",
    type=['csv'],
    label_visibility="collapsed"
)

if uploaded_file is not None:
    try:
        # CSVデータの読み込み（複数のエンコーディングを試行）
        encodings = ['utf-8-sig', 'utf-8', 'shift-jis', 'cp932', 'iso-2022-jp', 'euc-jp']
        df = None
        last_error = None
        
        for encoding in encodings:
            try:
                uploaded_file.seek(0)  # ファイルポインタを先頭に戻す
                df = pd.read_csv(uploaded_file, encoding=encoding, header=None, on_bad_lines='warn', engine='python')
                break
            except (UnicodeDecodeError, UnicodeError):
                last_error = encoding
                continue
        
        if df is None:
            raise ValueError(f"CSVファイルの読み込みに失敗しました。サポートされているエンコーディングで保存されているか確認してください。")
        
        # データプレビュー
        st.markdown('<div class="section-title">データプレビュー</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="info-box">{len(df)}行 × {len(df.columns)}列のデータが読み込まれました</div>', unsafe_allow_html=True)
        st.dataframe(df.head(10), use_container_width=True)
        
        # 変換ボタン
        st.markdown('<div class="section-title">変換してダウンロード</div>', unsafe_allow_html=True)
        
        if st.button("Excelに変換", use_container_width=True):
            with st.spinner('変換中...'):
                # テンプレートを読み込み
                wb = openpyxl.load_workbook(template_path)
                
                # "貼り付け用"シートを取得（存在しない場合は作成）
                sheet_name = "貼り付け用"
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # 既存の全てのデータをクリア
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.value = None
                else:
                    ws = wb.create_sheet(sheet_name)
                
                # A列の1行目からデータを書き込み（ヘッダーなし）
                for row_idx in range(len(df)):
                    for col_idx in range(len(df.columns)):
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=df.iloc[row_idx, col_idx])
                
                # メモリ上に保存
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                # 成功メッセージ
                st.markdown('<div class="success-box">変換が完了しました</div>', unsafe_allow_html=True)
                
                # ダウンロードボタン
                # アップロードされたファイル名から拡張子を除いて_convertedを追加
                original_name = uploaded_file.name.rsplit('.', 1)[0]
                download_filename = f"{original_name}_converted.xlsx"
                
                st.download_button(
                    label="ダウンロード",
                    data=output,
                    file_name=download_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
    except Exception as e:
        st.error(f"エラーが発生しました: {str(e)}")
        st.info("CSVファイルのエンコーディングや形式を確認してください")
else:
    # 使い方の説明
    st.markdown("""
    <div class="instruction-list">
        <div class="instruction-item">
            <div class="step-number">1</div>
            <div class="step-text">上のエリアにCSVファイルをアップロード</div>
        </div>
        <div class="instruction-item">
            <div class="step-number">2</div>
            <div class="step-text">データプレビューで内容を確認</div>
        </div>
        <div class="instruction-item">
            <div class="step-number">3</div>
            <div class="step-text">「Excelに変換」ボタンをクリック</div>
        </div>
        <div class="instruction-item">
            <div class="step-number">4</div>
            <div class="step-text">変換されたExcelファイルをダウンロード</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

# フッター
st.markdown('<hr>', unsafe_allow_html=True)
st.markdown('<div class="footer">Built with Streamlit</div>', unsafe_allow_html=True)
